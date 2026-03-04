"""
Donify - Excel Report Generator
--------------------------------
Lit metrics_results.csv 
Crée un fichier final metrics_results.xlsx
contenant :
- Tableau des métriques
- Graphes de comparaison
"""

import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment


# =========================
# 1) Chemins
# =========================

INPUT_CSV = r".\reports\metrics\metrics_results.csv"
OUTPUT_XLSX = r".\reports\metrics\metrics_results.xlsx"
TEMP_IMG = r".\reports\metrics\temp_chart.png"


# =========================
# 2) Charger les données
# =========================

if not os.path.exists(INPUT_CSV):
    raise FileNotFoundError(f"Fichier introuvable : {INPUT_CSV}")

# Lire avec le bon séparateur
df = pd.read_csv(INPUT_CSV, sep=";")

# Mettre la colonne modèle en index
df = df.set_index("Colonne1")

# Convertir en float
df = df.astype(float)

print("DataFrame chargé :")
print(df)

print("Shape:", df.shape)
print("Colonnes:", df.columns)
print("Dtypes:", df.dtypes)
print("Head:")
print(df.head())

# =========================
# 3) Créer graphique global
# =========================

plt.figure()
df.plot(kind="bar")
plt.title("Comparaison des modèles - Toutes les métriques")
plt.ylabel("Score")
plt.xticks(rotation=20)
plt.tight_layout()
plt.savefig(TEMP_IMG, dpi=200)
plt.close()

# =========================
# 4) Créer fichier Excel
# =========================

wb = Workbook()

# ---- Feuille 1 : Tableau
ws = wb.active
ws.title = "Metrics"

ws["A1"] = "Donify - Comparaison des modèles"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:F1")

df_to_write = df.copy()
df_to_write.insert(0, "Modèles", df_to_write.index)

for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), start=3):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="center")

# Ajuster largeur colonnes
ws.column_dimensions["A"].width = 32
for col in ["B", "C", "D", "E", "F"]:
    ws.column_dimensions[col].width = 14

# ---- Feuille 2 : Graphes
ws2 = wb.create_sheet("Graphs")

ws2["A1"] = "Graphiques de comparaison"
ws2["A1"].font = Font(bold=True, size=14)

img = XLImage(TEMP_IMG)
img.width = 900
ws2.add_image(img, "A3")

# Sauvegarder
wb.save(OUTPUT_XLSX)

# Supprimer image temporaire
os.remove(TEMP_IMG)

print("✅ Fichier final créé :")
print(OUTPUT_XLSX)